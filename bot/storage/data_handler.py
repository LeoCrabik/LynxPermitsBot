import json
import os
from openpyxl import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension
from keyboard import *
import re


def save_to_json(data, username, company_name):
    if not os.path.exists('data'):
        os.mkdir('data')
    file_path = f'data/user_data/{username}_{re.sub(r"[^A-Za-z0-9]+", "", company_name)}.json'

    with open(file_path, 'w') as file:
        json.dump(data, file, ensure_ascii=False, indent=4)


def read_json(filename):
    if not os.path.exists('data'):
        os.mkdir('data')
    file_path = f'data/user_data/{filename}'
    if os.path.isfile(file_path):
        with open(file_path) as file:
            data = json.load(file)
            print(data)
            return data

    empty_data = read_questions()
    return {item['var_name']: "" for item in empty_data}


# def read_questions(file_path=f'data/questions.json'):
#     if os.path.isfile(file_path):
#         with open(file_path) as file:
#             data = json.load(file)
#             return data
#     return False


def save_to_excel(data, username):
    if not os.path.exists('data'):
        os.mkdir('data')
    file_path = f'data/output/{username}.xlsx'
    names_data = read_questions()
    item_names = {item['var_name']: item['name'] for item in names_data}

    wb = Workbook()
    ws = wb.active
    i = 1
    for key, value in data.items():
        if key not in ['state', 'multimedia']:
            ws[f'A{i}'] = item_names[key]
            ws[f'B{i}'] = value if value else 'N/A'
            i += 1
    for column in ws.columns:
        max_length = 0
        column_name = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_name].width = adjusted_width
    wb.save(file_path)


def read_questions():
    questions_list = [
        {"var_name": "applicant_company", "name": "Applicant Co:"},
        {"var_name": "dot_number", "name": "DOT #:"},
        {"var_name": "date", "name": "Date:"},
        {"var_name": "contact_name", "name": "Contact:"},
        {"var_name": "contact_phone", "name": "Phone:"},
        {"var_name": "contact_email", "name": "Email:"},
        {"var_name": "federal_id", "name": "Federal ID #"},
        {"var_name": "permit_type", "name": "Permit Type:"},
        {"var_name": "application_load_number", "name": "Application Load/Pro #"},
        {"var_name": "start_address", "name": "Start St Address:"},
        {"var_name": "destination_address", "name": "Dest St Address:"},
        {"var_name": "tractor_number", "name": "Trct #:"},
        {"var_name": "tractor_year", "name": "Trct Year:"},
        {"var_name": "tractor_make", "name": "Trct Make:"},
        {"var_name": "tractor_license", "name": "Trct Lic:"},
        {"var_name": "tractor_state", "name": "Trct St:"},
        {"var_name": "tractor_sn", "name": "Trct SN:"},
        {"var_name": "tractor_axles_number", "name": "Trct Axles Number:"},
        {"var_name": "trailer_number", "name": "Trlr #:"},
        {"var_name": "trailer_year", "name": "Trlr Year:"},
        {"var_name": "trailer_dimension", "name": "Trlr Dmsn:"},
        {"var_name": "trailer_type", "name": "Trlr Type:"},
        {"var_name": "trailer_make", "name": "Trlr Make:"},
        {"var_name": "trailer_license", "name": "Trlr Lic:"},
        {"var_name": "trailer_state", "name": "Trlr St:"},
        {"var_name": "trailer_sn", "name": "Trlr SN:"},
        {"var_name": "trailer_axles_number", "name": "Trlr Axles Number:"},
        {"var_name": "load_description", "name": "Description of Load:"},
        {"var_name": "number_of_pieces", "name": "No. Pieces:"},
        {"var_name": "load_machinery_make", "name": "Machinery Make:"},
        {"var_name": "load_model", "name": "Model:"},
        {"var_name": "load_sn", "name": "SN(s):"},
        {"var_name": "load_weight", "name": "Load Weight:"},
        {"var_name": "load_height", "name": "Load Height:"},
        {"var_name": "load_width", "name": "Load Width:"},
        {"var_name": "load_length", "name": "Load Length:"},
        {"var_name": "total_weight", "name": "Total Weight:"},
        {"var_name": "total_height", "name": "Total Height:"},
        {"var_name": "total_width", "name": "Total Width:"},
        {"var_name": "total_length", "name": "Total Length:"},
        {"var_name": "overhang_front", "name": "Overhang Front:"},
        {"var_name": "overhang_rear", "name": "Overhang Rear:"},
        {"var_name": "how_is_it_loaded", "name": "How Is It Loaded:"},
        {"var_name": "axle1_weight", "name": "Axle №1 Weight:"},
        {"var_name": "axle1_spacings", "name": "Axle №1 Spacings:"},
        {"var_name": "axle2_weight", "name": "Axle №2 Weight:"},
        {"var_name": "axle2_spacings", "name": "Axle №2 Spacings:"},
        {"var_name": "axle3_weight", "name": "Axle №3 Weight:"},
        {"var_name": "axle3_spacings", "name": "Axle №3 Spacings:"},
        {"var_name": "axle4_weight", "name": "Axle №4 Weight:"},
        {"var_name": "axle4_spacings", "name": "Axle №4 Spacings:"},
        {"var_name": "axle5_weight", "name": "Axle №5 Weight:"},
        {"var_name": "axle5_spacings", "name": "Axle №5 Spacings:"},
        {"var_name": "axle6_weight", "name": "Axle №6 Weight:"},
        {"var_name": "axle6_spacings", "name": "Axle №6 Spacings:"},
        {"var_name": "axle7_weight", "name": "Axle №7 Weight:"},
        {"var_name": "axle7_spacings", "name": "Axle №7 Spacings:"},
        {"var_name": "axle8_weight", "name": "Axle №8 Weight:"},
        {"var_name": "axle8_spacings", "name": "Axle №8 Spacings:"},
        {"var_name": "axle9_weight", "name": "Axle №9 Weight:"},
        {"var_name": "axle9_spacings", "name": "Axle №9 Spacings:"},
        {"var_name": "axle10_weight", "name": "Axle №10 Weight:"},
        {"var_name": "axle10_spacings", "name": "Axle №10 Spacings:"},
        {"var_name": "axle11_weight", "name": "Axle №11 Weight:"},
        {"var_name": "axle11_spacings", "name": "Axle №11 Spacings:"},
        {"var_name": "axle12_weight", "name": "Axle №12 Weight:"},
        {"var_name": "comment", "name": "Comment:"}
    ]
    return questions_list


def find_jsons(username):
    folder_path = f'data/user_data/'
    matching_files = []
    for filename in os.listdir(folder_path):
        if filename.startswith(f'{username}'):
            matching_files.append(filename)
    return matching_files


if __name__ == '__main__':
    main_data = (read_questions(f'C:/Users/leocr/Desktop/LynxPermitsBot/bot/data/questions.json'))
    print(main_data)
    data_list = {main_data['states'][item]['var_name']: "" for item in main_data['states']}
    print(data_list)
